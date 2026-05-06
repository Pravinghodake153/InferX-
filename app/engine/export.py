"""
Export Module - Generate audit-ready PDF, Excel, and JSON reports.

Supports:
  - PDF Report: Full evaluation with verdicts, evidence, confidence
  - Excel Matrix: Structured spreadsheet for analysis
  - Audit JSON: SHA-256 hashed tamper-proof log
"""
import hashlib
import json
import os
import uuid
from datetime import datetime
from typing import Dict, Any, List, Optional

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm, cm
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    HRFlowable, PageBreak
)
from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# ═══════════════════════════════════════════
#  AUDIT METADATA
# ═══════════════════════════════════════════

def generate_evaluation_id() -> str:
    """Generate a unique evaluation ID."""
    return f"EVAL-{uuid.uuid4().hex[:8].upper()}"


def generate_audit_hash(data: Dict[str, Any]) -> str:
    """Generate SHA-256 hash of evaluation data for tamper detection."""
    canonical = json.dumps(data, sort_keys=True, default=str)
    return hashlib.sha256(canonical.encode()).hexdigest()


def build_audit_metadata(data: Dict[str, Any]) -> Dict[str, Any]:
    """Build audit metadata envelope for the evaluation."""
    eval_id = generate_evaluation_id()
    timestamp = datetime.now().isoformat()
    sha_hash = generate_audit_hash(data)

    return {
        "evaluation_id": eval_id,
        "timestamp": timestamp,
        "model_version": "InferX v2.0",
        "pipeline_version": "7-step",
        "provider": data.get("provider", "unknown"),
        "sha256_hash": sha_hash,
        "criteria_count": len(data.get("criteria", [])),
        "evidence_count": len(data.get("evidence", [])),
        "issues_count": len(data.get("issues", [])),
    }


# ═══════════════════════════════════════════
#  PDF REPORT
# ═══════════════════════════════════════════

# Color palette
DARK_BG = colors.HexColor("#0f1729")
HEADER_BG = colors.HexColor("#1a2744")
PASS_COLOR = colors.HexColor("#22c55e")
FAIL_COLOR = colors.HexColor("#ef4444")
REVIEW_COLOR = colors.HexColor("#f59e0b")
ACCENT_BLUE = colors.HexColor("#3b82f6")
TEXT_WHITE = colors.white
TEXT_GRAY = colors.HexColor("#94a3b8")
ROW_ALT = colors.HexColor("#1e293b")


def generate_pdf_report(data: Dict[str, Any], output_path: str) -> str:
    """Generate a professional PDF evaluation report."""
    doc = SimpleDocTemplate(
        output_path,
        pagesize=A4,
        rightMargin=15 * mm,
        leftMargin=15 * mm,
        topMargin=20 * mm,
        bottomMargin=15 * mm
    )

    styles = getSampleStyleSheet()

    # Custom styles
    title_style = ParagraphStyle(
        'CustomTitle', parent=styles['Title'],
        fontSize=22, textColor=ACCENT_BLUE,
        spaceAfter=6, alignment=TA_CENTER
    )
    subtitle_style = ParagraphStyle(
        'CustomSubtitle', parent=styles['Normal'],
        fontSize=10, textColor=TEXT_GRAY,
        alignment=TA_CENTER, spaceAfter=12
    )
    section_style = ParagraphStyle(
        'SectionHeader', parent=styles['Heading2'],
        fontSize=14, textColor=ACCENT_BLUE,
        spaceBefore=16, spaceAfter=8
    )
    body_style = ParagraphStyle(
        'Body', parent=styles['Normal'],
        fontSize=11, textColor=colors.black,
        spaceAfter=4
    )
    small_style = ParagraphStyle(
        'Small', parent=styles['Normal'],
        fontSize=10, textColor=TEXT_GRAY,
        spaceAfter=2
    )

    content = []
    audit = build_audit_metadata(data)

    # ── Title ──
    content.append(Paragraph("InferX - AI Tender Evaluation Report", title_style))
    content.append(Paragraph("Government-Grade Document Intelligence System", subtitle_style))
    content.append(Spacer(1, 4 * mm))

    # ── Audit Header ──
    audit_info = [
        ["Evaluation ID", audit["evaluation_id"], "Timestamp", audit["timestamp"][:19]],
        ["Provider", audit["provider"].upper(), "Pipeline", audit["pipeline_version"]],
        ["SHA-256", audit["sha256_hash"][:32] + "...", "Model", audit["model_version"]],
    ]
    audit_table = Table(audit_info, colWidths=[80, 150, 80, 150])
    audit_table.setStyle(TableStyle([
        ('FONTSIZE', (0, 0), (-1, -1), 12),
        ('TEXTCOLOR', (0, 0), (0, -1), ACCENT_BLUE),
        ('TEXTCOLOR', (2, 0), (2, -1), ACCENT_BLUE),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (2, 0), (2, -1), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor("#f8fafc")),
    ]))
    content.append(audit_table)
    content.append(Spacer(1, 6 * mm))

    # ── Summary (mandatory-aware verdict) ──
    evals = data.get("evaluation", [])
    criteria_lookup = {c.get("criterion_id"): c for c in data.get("criteria", [])}

    # Separate mandatory vs optional results
    mandatory_fails = 0
    mandatory_reviews = 0
    optional_fails = 0

    pass_count = sum(1 for e in evals if e.get("result") == "PASS")
    fail_count = sum(1 for e in evals if e.get("result") == "FAIL")
    review_count = sum(1 for e in evals if e.get("result") == "REVIEW")

    for e in evals:
        cid = e.get("criterion_id") or e.get("criteria_id")
        criterion = criteria_lookup.get(cid, {})
        is_mandatory = criterion.get("mandatory", True)  # Default to mandatory if unknown

        if e.get("result") == "FAIL":
            if is_mandatory:
                mandatory_fails += 1
            else:
                optional_fails += 1
        elif e.get("result") == "REVIEW":
            if is_mandatory:
                mandatory_reviews += 1

    # Overall verdict based ONLY on mandatory criteria
    if mandatory_fails > 0:
        verdict_text = "NOT ELIGIBLE"
        verdict_color = FAIL_COLOR
    elif mandatory_reviews > 0:
        verdict_text = "REVIEW REQUIRED"
        verdict_color = REVIEW_COLOR
    else:
        verdict_text = "ELIGIBLE"
        verdict_color = PASS_COLOR

    verdict_style = ParagraphStyle(
        'Verdict', parent=styles['Title'],
        fontSize=20, textColor=verdict_color,
        alignment=TA_CENTER, spaceBefore=4, spaceAfter=4
    )
    content.append(Paragraph(f"Final Verdict: {verdict_text}", verdict_style))

    # Show optional fail count as info if any
    if optional_fails > 0:
        opt_note = ParagraphStyle('OptNote', parent=styles['Normal'],
                                   fontSize=9, textColor=REVIEW_COLOR, alignment=TA_CENTER)
        content.append(Paragraph(
            f"Note: {optional_fails} optional criteria failed (does not affect eligibility)",
            opt_note
        ))

    summary_data = [
        ["Total Criteria", "Passed", "Failed (Mandatory)", "Review"],
        [str(len(evals)), str(pass_count), str(mandatory_fails), str(review_count)]
    ]
    summary_table = Table(summary_data, colWidths=[115, 115, 115, 115])
    summary_table.setStyle(TableStyle([
        ('FONTSIZE', (0, 0), (-1, -1), 13),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('BACKGROUND', (0, 0), (-1, 0), HEADER_BG),
        ('TEXTCOLOR', (0, 0), (-1, 0), TEXT_WHITE),
        ('TEXTCOLOR', (1, 1), (1, 1), PASS_COLOR),
        ('TEXTCOLOR', (2, 1), (2, 1), FAIL_COLOR),
        ('TEXTCOLOR', (3, 1), (3, 1), REVIEW_COLOR),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#334155")),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
    ]))
    content.append(Spacer(1, 4 * mm))
    content.append(summary_table)
    content.append(Spacer(1, 6 * mm))

    # ── Evaluation Results ──
    content.append(HRFlowable(width="100%", color=ACCENT_BLUE, thickness=1))
    content.append(Paragraph("Evaluation Results", section_style))

    if evals:
        eval_header = ["ID", "Criterion", "Type", "Required", "Found", "Verdict"]
        eval_rows = [eval_header]
        for e in evals:
            verdict = e.get("result", "")
            cid = e.get("criteria_id", "")
            criterion = criteria_lookup.get(cid, {})
            c_type = "Mandatory" if criterion.get("mandatory", True) else "Optional"
            
            eval_rows.append([
                cid,
                Paragraph(e.get("criteria_name", ""), body_style),
                c_type,
                e.get("required_value", "-"),
                e.get("extracted_value", "-"),
                verdict
            ])

        eval_table = Table(eval_rows, colWidths=[40, 130, 60, 80, 90, 60])
        eval_table.setStyle(TableStyle([
            ('FONTSIZE', (0, 0), (-1, -1), 12),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 0), (-1, 0), HEADER_BG),
            ('TEXTCOLOR', (0, 0), (-1, 0), TEXT_WHITE),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#334155")),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#f1f5f9")]),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))

        # Color verdict cells
        for i, e in enumerate(evals, 1):
            v = e.get("result", "")
            if v == "PASS":
                eval_table.setStyle(TableStyle([('TEXTCOLOR', (5, i), (5, i), PASS_COLOR)]))
            elif v == "FAIL":
                eval_table.setStyle(TableStyle([('TEXTCOLOR', (5, i), (5, i), FAIL_COLOR)]))
            elif v == "REVIEW":
                eval_table.setStyle(TableStyle([('TEXTCOLOR', (5, i), (5, i), REVIEW_COLOR)]))

        content.append(eval_table)

    # ── Reasoning ──
    content.append(Spacer(1, 6 * mm))
    content.append(Paragraph("Detailed Reasoning", section_style))
    for e in evals:
        reason = e.get("reason", "No reason provided")
        name = e.get("criteria_name", e.get("criteria_id", ""))
        verdict = e.get("result", "")
        icon = "PASS" if verdict == "PASS" else ("FAIL" if verdict == "FAIL" else "REVIEW")
        content.append(Paragraph(
            f"<b>{e.get('criteria_id', '')} - {name}</b> [{icon}]", body_style
        ))
        content.append(Paragraph(f"→ {reason}", small_style))
        content.append(Spacer(1, 2 * mm))

    # ── Vigilance ──
    issues = data.get("issues", [])
    if issues:
        content.append(Spacer(1, 4 * mm))
        content.append(HRFlowable(width="100%", color=FAIL_COLOR, thickness=1))
        content.append(Paragraph("Vigilance Alerts", section_style))
        for issue in issues:
            sev = issue.get("severity", "")
            content.append(Paragraph(
                f"<b>[{sev}]</b> {issue.get('issue_type', '')}: {issue.get('reason', '')}",
                body_style
            ))

    # ── Verification ──
    # Support both singular and plural keys for verification data
    verifications = data.get("verification", []) + data.get("verifications", [])
    if verifications:
        content.append(Spacer(1, 4 * mm))
        content.append(HRFlowable(width="100%", color=ACCENT_BLUE, thickness=1))
        content.append(Paragraph("Identifier Verification", section_style))
        for v in verifications:
            status = v.get("status", "")
            content.append(Paragraph(
                f"<b>{v.get('identifier_type', '')}</b>: {v.get('identifier', '')} - {status}",
                body_style
            ))
            if v.get("details"):
                content.append(Paragraph(f"  {v['details']}", small_style))  # details retained

    # ── Procurement Officer Sign-Off ──
    content.append(Spacer(1, 12 * mm))
    content.append(HRFlowable(width="100%", color=ACCENT_BLUE, thickness=1.5))
    content.append(Paragraph("Authorized Sign-Off", section_style))
    content.append(Spacer(1, 4 * mm))

    signoff_style = ParagraphStyle(
        'SignOff', parent=styles['Normal'],
        fontSize=10, textColor=colors.black,
        spaceAfter=6, leading=18
    )
    content.append(Paragraph("I have reviewed the above evaluation report and confirm the findings.", signoff_style))
    content.append(Spacer(1, 6 * mm))

    signoff_data = [
        ["Procurement Officer Name:", "______________________________"],
        ["Designation:", "______________________________"],
        ["Signature:", "______________________________"],
        ["Date:", "______________________________"],
        ["Audit Hash Verified:", audit['sha256_hash'][:32] + "..."],
    ]
    signoff_table = Table(signoff_data, colWidths=[150, 310])
    signoff_table.setStyle(TableStyle([
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('TEXTCOLOR', (0, 0), (0, -1), ACCENT_BLUE),
        ('TEXTCOLOR', (1, -1), (1, -1), TEXT_GRAY),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('LINEBELOW', (0, 0), (-1, -2), 0.5, colors.HexColor("#e2e8f0")),
    ]))
    content.append(signoff_table)

    # ── Footer ──
    content.append(Spacer(1, 10 * mm))
    content.append(HRFlowable(width="100%", color=TEXT_GRAY, thickness=0.5))
    content.append(Paragraph(
        f"Generated by InferX v2.0 | {audit['timestamp'][:19]} | SHA-256: {audit['sha256_hash'][:16]}...",
        ParagraphStyle('Footer', parent=styles['Normal'], fontSize=9, textColor=TEXT_GRAY, alignment=TA_CENTER)
    ))
    content.append(Paragraph(
        "This is a machine-generated report. All evaluations require human verification before final decisions.",
        ParagraphStyle('Disclaimer', parent=styles['Normal'], fontSize=9, textColor=REVIEW_COLOR,
                       alignment=TA_CENTER, spaceBefore=4)
    ))

    doc.build(content)
    return output_path


# ═══════════════════════════════════════════
#  EXCEL MATRIX
# ═══════════════════════════════════════════

# Color fills
HEADER_FILL = PatternFill(start_color="1a2744", end_color="1a2744", fill_type="solid")
PASS_FILL = PatternFill(start_color="dcfce7", end_color="dcfce7", fill_type="solid")
FAIL_FILL = PatternFill(start_color="fecaca", end_color="fecaca", fill_type="solid")
REVIEW_FILL = PatternFill(start_color="fef3c7", end_color="fef3c7", fill_type="solid")
ALT_ROW_FILL = PatternFill(start_color="f1f5f9", end_color="f1f5f9", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=12)
BODY_FONT = Font(size=11, color="1e293b")
THIN_BORDER = Border(
    left=Side(style='thin', color='cbd5e1'),
    right=Side(style='thin', color='cbd5e1'),
    top=Side(style='thin', color='cbd5e1'),
    bottom=Side(style='thin', color='cbd5e1'),
)


def generate_excel_report(data: Dict[str, Any], output_path: str) -> str:
    """Generate an Excel evaluation matrix.
    
    Handles both:
    - Single-bidder format: data["evaluation"] = [...]
    - Multi-bidder consolidated format: data["bidder_results"] = [{ "evaluation": [...] }, ...]
    """
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    audit = build_audit_metadata(data)
    criteria_lookup = {c.get("criterion_id"): c for c in data.get("criteria", [])}

    # Resolve evaluation data: flatten bidder_results if present
    bidder_results = data.get("bidder_results", [])
    flat_evals = data.get("evaluation", [])

    # If consolidated format (multi-bidder), build per-bidder sheets
    if bidder_results and len(bidder_results) > 0:
        # ── Sheet 1: Summary ──
        ws_summary = wb.active
        ws_summary.title = "Summary"

        ws_summary.append(["InferX -- Consolidated Evaluation Report"])
        ws_summary.merge_cells('A1:F1')
        ws_summary['A1'].font = Font(bold=True, size=14, color="3b82f6")

        ws_summary.append([f"Evaluation ID: {audit['evaluation_id']}", "", "",
                   f"Timestamp: {audit['timestamp'][:19]}"])
        ws_summary.append([f"Provider: {audit['provider']}", "", "",
                   f"SHA-256: {audit['sha256_hash'][:32]}..."])
        ws_summary.append([f"Total Bidders: {len(bidder_results)}"])
        ws_summary.append([])

        # Summary table header
        summary_headers = ["#", "Bidder Name", "Verdict", "Pass", "Fail", "Review"]
        ws_summary.append(summary_headers)
        summary_header_row = ws_summary.max_row
        for col_idx, _ in enumerate(summary_headers, 1):
            cell = ws_summary.cell(row=summary_header_row, column=col_idx)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal='center')
            cell.border = THIN_BORDER

        for idx, br in enumerate(bidder_results, 1):
            bidder_name = br.get("bidder_name", f"Bidder {idx}")
            verdict = br.get("verdict", "REVIEW_REQUIRED")
            evals = br.get("evaluation", [])
            pass_count = sum(1 for e in evals if e.get("result") == "PASS")
            fail_count = sum(1 for e in evals if e.get("result") == "FAIL")
            review_count = sum(1 for e in evals if e.get("result") == "REVIEW")

            ws_summary.append([idx, bidder_name, verdict, pass_count, fail_count, review_count])
            current_row = ws_summary.max_row

            # Color the verdict cell
            verdict_cell = ws_summary.cell(row=current_row, column=3)
            if verdict == "ELIGIBLE":
                verdict_cell.fill = PASS_FILL
                verdict_cell.font = Font(bold=True, color="166534", size=11)
            elif verdict == "NOT_ELIGIBLE":
                verdict_cell.fill = FAIL_FILL
                verdict_cell.font = Font(bold=True, color="991b1b", size=11)
            else:
                verdict_cell.fill = REVIEW_FILL
                verdict_cell.font = Font(bold=True, color="92400e", size=11)

            for col_idx in range(1, len(summary_headers) + 1):
                cell = ws_summary.cell(row=current_row, column=col_idx)
                cell.border = THIN_BORDER
                if not cell.font or not cell.font.bold:
                    cell.font = BODY_FONT
                if idx % 2 == 0:
                    if col_idx != 3:  # Don't override verdict color
                        cell.fill = ALT_ROW_FILL

        summary_col_widths = [6, 30, 18, 10, 10, 10]
        for i, w in enumerate(summary_col_widths, 1):
            ws_summary.column_dimensions[get_column_letter(i)].width = w

        # ── Per-Bidder Sheets ──
        for idx, br in enumerate(bidder_results, 1):
            bidder_name = br.get("bidder_name", f"Bidder {idx}")
            # Sheet names max 31 chars, no special chars
            safe_name = bidder_name[:28].replace("/", "-").replace("\\", "-").replace("*", "").replace("?", "").replace("[", "").replace("]", "").replace(":", "-")
            ws = wb.create_sheet(title=safe_name)

            ws.append([f"Bidder: {bidder_name}"])
            ws.merge_cells('A1:I1')
            ws['A1'].font = Font(bold=True, size=14, color="3b82f6")

            verdict = br.get("verdict", "REVIEW_REQUIRED")
            ws.append([f"Verdict: {verdict}"])
            ws['A2'].font = Font(bold=True, size=12,
                                 color="166534" if verdict == "ELIGIBLE" else ("991b1b" if verdict == "NOT_ELIGIBLE" else "92400e"))
            ws.append([])

            # Headers
            headers = ["ID", "Criterion", "Category", "Type", "Required", "Found", "Confidence", "Verdict", "Reason"]
            ws.append(headers)
            header_row = ws.max_row
            for col_idx, _ in enumerate(headers, 1):
                cell = ws.cell(row=header_row, column=col_idx)
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.alignment = Alignment(horizontal='center')
                cell.border = THIN_BORDER

            evals = br.get("evaluation", [])
            for i, e in enumerate(evals):
                cid = e.get("criterion_id", e.get("criteria_id", ""))
                criterion = criteria_lookup.get(cid, {})
                c_type = "Mandatory" if criterion.get("mandatory", e.get("mandatory", True)) else "Optional"

                row = [
                    cid,
                    e.get("criteria_name", ""),
                    e.get("category", ""),
                    c_type,
                    str(e.get("required_value", "-")),
                    str(e.get("evidence_found", e.get("extracted_value", "-"))),
                    str(e.get("confidence", "")),
                    e.get("result", ""),
                    str(e.get("reason", "")),
                ]
                ws.append(row)
                current_row = ws.max_row

                result = e.get("result", "")
                for col_idx in range(1, len(headers) + 1):
                    cell = ws.cell(row=current_row, column=col_idx)
                    cell.font = BODY_FONT
                    cell.border = THIN_BORDER
                    if i % 2 == 1:
                        cell.fill = ALT_ROW_FILL

                # Color verdict cell
                verdict_cell = ws.cell(row=current_row, column=8)
                if result == "PASS":
                    verdict_cell.fill = PASS_FILL
                    verdict_cell.font = Font(bold=True, color="166534", size=11)
                elif result == "FAIL":
                    verdict_cell.fill = FAIL_FILL
                    verdict_cell.font = Font(bold=True, color="991b1b", size=11)
                elif result == "REVIEW":
                    verdict_cell.fill = REVIEW_FILL
                    verdict_cell.font = Font(bold=True, color="92400e", size=11)

            col_widths = [8, 25, 12, 12, 15, 15, 12, 12, 50]
            for i, w in enumerate(col_widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = w

    else:
        # ── Single-bidder fallback (legacy format) ──
        ws = wb.active
        ws.title = "Evaluation Matrix"

        ws.append(["InferX -- AI Tender Evaluation Report"])
        ws.merge_cells('A1:G1')
        ws['A1'].font = Font(bold=True, size=14, color="3b82f6")

        ws.append([f"Evaluation ID: {audit['evaluation_id']}", "", "",
                   f"Timestamp: {audit['timestamp'][:19]}"])
        ws.append([f"Provider: {audit['provider']}", "", "",
                   f"SHA-256: {audit['sha256_hash'][:32]}..."])
        ws.append([])

        headers = ["ID", "Criterion", "Category", "Type", "Required", "Found", "Confidence", "Verdict", "Reason"]
        ws.append(headers)
        header_row = ws.max_row
        for col_idx, _ in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_idx)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal='center')
            cell.border = THIN_BORDER

        for i, e in enumerate(flat_evals):
            cid = e.get("criteria_id", "")
            criterion = criteria_lookup.get(cid, {})
            c_type = "Mandatory" if criterion.get("mandatory", True) else "Optional"

            row = [
                cid,
                e.get("criteria_name", ""),
                e.get("category", ""),
                c_type,
                str(e.get("required_value", "-")),
                str(e.get("extracted_value", "-")),
                str(e.get("confidence", "")),
                e.get("result", ""),
                str(e.get("reason", "")),
            ]
            ws.append(row)
            current_row = ws.max_row

            verdict = e.get("result", "")
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.font = BODY_FONT
                cell.border = THIN_BORDER
                if i % 2 == 1:
                    cell.fill = ALT_ROW_FILL

            verdict_cell = ws.cell(row=current_row, column=8)
            if verdict == "PASS":
                verdict_cell.fill = PASS_FILL
                verdict_cell.font = Font(bold=True, color="166534", size=11)
            elif verdict == "FAIL":
                verdict_cell.fill = FAIL_FILL
                verdict_cell.font = Font(bold=True, color="991b1b", size=11)
            elif verdict == "REVIEW":
                verdict_cell.fill = REVIEW_FILL
                verdict_cell.font = Font(bold=True, color="92400e", size=11)

        col_widths = [8, 25, 12, 12, 15, 15, 12, 12, 50]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # ── Vigilance Sheet ──
    ws_vig = wb.create_sheet("Vigilance")
    ws_vig.append(["Severity", "Type", "Criterion", "Reason", "Confidence"])
    for col_idx in range(1, 6):
        cell = ws_vig.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    # Collect issues from all bidder_results or top-level
    all_issues = data.get("issues", [])
    for br in bidder_results:
        all_issues.extend(br.get("issues", []))
    for issue in all_issues:
        ws_vig.append([
            issue.get("severity", ""),
            issue.get("issue_type", ""),
            issue.get("criterion_id", ""),
            issue.get("reason", ""),
            issue.get("confidence", ""),
        ])

    # ── Verification Sheet ──
    ws_ver = wb.create_sheet("Verification")
    ws_ver.append(["Identifier", "Type", "Status", "Confidence", "Details"])
    for col_idx in range(1, 6):
        cell = ws_ver.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    # Aggregate verification data from both possible keys
    all_verifications = data.get("verification", []) + data.get("verifications", [])
    for br in bidder_results:
        all_verifications.extend(br.get("verification", []))
        all_verifications.extend(br.get("verifications", []))
    for v in all_verifications:
        ws_ver.append([
            v.get("identifier", ""),
            v.get("identifier_type", ""),
            v.get("status", ""),
            v.get("confidence", ""),
            v.get("details", ""),
        ])

    wb.save(output_path)
    return output_path


# ═══════════════════════════════════════════
#  AUDIT JSON LOG
# ═══════════════════════════════════════════

def generate_audit_json(data: Dict[str, Any]) -> Dict[str, Any]:
    """
    Generate a tamper-proof audit JSON log with SHA-256 hash.
    """
    audit = build_audit_metadata(data)

    return {
        "audit_metadata": audit,
        "evaluation_data": {
            "criteria": data.get("criteria", []),
            "evidence": data.get("evidence", []),
            "evaluation": data.get("evaluation", []),
            "final_evaluation": data.get("final_evaluation", []),
            "field_mappings": data.get("field_mappings", []),
            "verification": data.get("verification", []),
            "pii_masking": data.get("pii_masking", {}),
            "issues": data.get("issues", []),
            "pipeline_steps": data.get("pipeline_steps", []),
        },
        "integrity": {
            "algorithm": "SHA-256",
            "hash": audit["sha256_hash"],
            "tamper_proof": True,
            "generated_at": audit["timestamp"],
            "disclaimer": "This audit log is machine-generated. Verify hash for tamper detection.",
        }
    }


# ═══════════════════════════════════════════
#  CONSOLIDATED MULTI-BIDDER PDF
# ═══════════════════════════════════════════

def generate_consolidated_pdf(data: Dict[str, Any], output_path: str) -> str:
    """
    Generate a consolidated multi-bidder comparison PDF report.
    
    Expected data format:
    {
        "criteria": [...],
        "bidder_results": [
            { "bidder_name": "...", "verdict": "...", "evaluation": [...], ... },
            ...
        ],
        "summary": { "eligible": N, "not_eligible": N, "review_required": N }
    }
    """
    from reportlab.lib.pagesizes import landscape
    
    styles = getSampleStyleSheet()
    
    page_size = landscape(A4)
    doc = SimpleDocTemplate(
        output_path, pagesize=page_size,
        topMargin=15 * mm, bottomMargin=15 * mm,
        leftMargin=15 * mm, rightMargin=15 * mm
    )
    content = []

    # Styles
    title_style = ParagraphStyle('ConsoTitle', parent=styles['Title'],
                                  fontSize=18, textColor=ACCENT_BLUE,
                                  alignment=TA_CENTER, spaceAfter=4)
    section_style = ParagraphStyle('ConsoSection', parent=styles['Heading2'],
                                    fontSize=13, textColor=ACCENT_BLUE,
                                    spaceBefore=8, spaceAfter=6)
    body_style = ParagraphStyle('ConsoBody', parent=styles['Normal'],
                                 fontSize=11, textColor=colors.black,
                                 spaceAfter=3, leading=12)
    small_style = ParagraphStyle('ConsoSmall', parent=styles['Normal'],
                                  fontSize=9, textColor=TEXT_GRAY, leading=11)

    # ── Title ──
    content.append(Paragraph("InferX - Consolidated Bidder Evaluation Report", title_style))
    content.append(Paragraph("Government Procurement • Multi-Bidder Comparison", small_style))
    content.append(Spacer(1, 4 * mm))

    # ── Audit header ──
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S UTC")
    eval_id = f"CONSOL-{uuid.uuid4().hex[:8].upper()}"
    hash_input = json.dumps(data, sort_keys=True, default=str).encode()
    sha256 = hashlib.sha256(hash_input).hexdigest()

    audit_data = [
        ["Evaluation ID", eval_id],
        ["Timestamp", timestamp],
        ["Total Bidders", str(data.get("bidder_count", len(data.get("bidder_results", []))))],
        ["SHA-256", sha256[:32] + "..."],
    ]
    audit_table = Table(audit_data, colWidths=[120, 350])
    audit_table.setStyle(TableStyle([
        ('FONTSIZE', (0, 0), (-1, -1), 12),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('TEXTCOLOR', (0, 0), (0, -1), ACCENT_BLUE),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
    ]))
    content.append(audit_table)
    content.append(Spacer(1, 6 * mm))

    # ── Executive Summary ──
    content.append(HRFlowable(width="100%", color=ACCENT_BLUE, thickness=1))
    content.append(Paragraph("Executive Summary", section_style))

    summary = data.get("summary", {})
    bidder_results = data.get("bidder_results", [])

    summary_data = [
        ["Total Bidders", "Eligible", "Not Eligible", "Review Required"],
        [str(len(bidder_results)),
         str(summary.get("eligible", 0)),
         str(summary.get("not_eligible", 0)),
         str(summary.get("review_required", 0))]
    ]
    summary_table = Table(summary_data, colWidths=[130, 130, 130, 130])
    summary_table.setStyle(TableStyle([
        ('FONTSIZE', (0, 0), (-1, -1), 13),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('BACKGROUND', (0, 0), (-1, 0), HEADER_BG),
        ('TEXTCOLOR', (0, 0), (-1, 0), TEXT_WHITE),
        ('TEXTCOLOR', (1, 1), (1, 1), PASS_COLOR),
        ('TEXTCOLOR', (2, 1), (2, 1), FAIL_COLOR),
        ('TEXTCOLOR', (3, 1), (3, 1), REVIEW_COLOR),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#334155")),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
    ]))
    content.append(summary_table)
    content.append(Spacer(1, 6 * mm))

    # ── Bidder Verdict Summary ──
    content.append(Paragraph("Bidder Verdict Summary", section_style))
    bidder_summary_header = ["#", "Bidder Name", "Verdict", "Pass", "Fail", "Review"]
    bidder_summary_rows = [bidder_summary_header]
    for idx, b in enumerate(bidder_results, 1):
        verdict = b.get("verdict", "UNKNOWN").replace("_", " ")
        bidder_summary_rows.append([
            str(idx),
            b.get("bidder_name", f"Bidder {idx}")[:30],
            verdict,
            str(b.get("pass_count", 0)),
            str(b.get("fail_count", 0)),
            str(b.get("review_count", 0)),
        ])

    bidder_table = Table(bidder_summary_rows, colWidths=[30, 180, 120, 60, 60, 60])
    table_styles = [
        ('FONTSIZE', (0, 0), (-1, -1), 12),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('ALIGN', (1, 0), (1, -1), 'LEFT'),
        ('BACKGROUND', (0, 0), (-1, 0), HEADER_BG),
        ('TEXTCOLOR', (0, 0), (-1, 0), TEXT_WHITE),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#e2e8f0")),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
    ]
    # Color-code verdict cells
    for row_idx, b in enumerate(bidder_results, 1):
        verdict = b.get("verdict", "")
        if verdict == "ELIGIBLE":
            table_styles.append(('TEXTCOLOR', (2, row_idx), (2, row_idx), PASS_COLOR))
            table_styles.append(('FONTNAME', (2, row_idx), (2, row_idx), 'Helvetica-Bold'))
        elif verdict == "NOT_ELIGIBLE":
            table_styles.append(('TEXTCOLOR', (2, row_idx), (2, row_idx), FAIL_COLOR))
            table_styles.append(('FONTNAME', (2, row_idx), (2, row_idx), 'Helvetica-Bold'))
        else:
            table_styles.append(('TEXTCOLOR', (2, row_idx), (2, row_idx), REVIEW_COLOR))
            table_styles.append(('FONTNAME', (2, row_idx), (2, row_idx), 'Helvetica-Bold'))

    bidder_table.setStyle(TableStyle(table_styles))
    content.append(bidder_table)
    content.append(Spacer(1, 8 * mm))

    # ── Criterion-wise Comparison Matrix ──
    criteria = data.get("criteria", [])
    if criteria and bidder_results:
        content.append(PageBreak())
        content.append(Paragraph("Criterion-wise Comparison Matrix", section_style))
        content.append(Paragraph(
            "Each cell shows the verdict for a bidder against each criterion.",
            small_style
        ))
        content.append(Spacer(1, 3 * mm))

        # Build header row
        header = ["Criterion"]
        for b in bidder_results:
            header.append(Paragraph(f"<b>{b.get('bidder_name', '?')[:15]}</b>", small_style))

        matrix_rows = [header]
        for c in criteria:
            row = [Paragraph(f"<b>{c.get('name', '?')[:35]}</b>", small_style)]
            for b in bidder_results:
                ev = next((e for e in b.get("evaluation", [])
                           if e.get("criterion_id") == c.get("criterion_id")), None)
                if ev:
                    result = ev.get("result", "-")
                    symbol = "[PASS]" if result == "PASS" else ("[FAIL]" if result == "FAIL" else "[REVIEW]")
                    row.append(Paragraph(f"{symbol}", body_style))
                else:
                    row.append(Paragraph("-", body_style))
            matrix_rows.append(row)

        n_bidders = len(bidder_results)
        col_w = min(80, int(500 / max(n_bidders, 1)))
        matrix_table = Table(matrix_rows, colWidths=[180] + [col_w] * n_bidders)
        matrix_table.setStyle(TableStyle([
            ('FONTSIZE', (0, 0), (-1, -1), 7),
            ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
            ('BACKGROUND', (0, 0), (-1, 0), HEADER_BG),
            ('TEXTCOLOR', (0, 0), (-1, 0), TEXT_WHITE),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#e2e8f0")),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
        ]))
        content.append(matrix_table)

    # ── Detailed Bidder Explanations ──
    content.append(PageBreak())
    content.append(Paragraph("Detailed Bidder Explanations", section_style))
    content.append(Paragraph("A to Z breakdown of each bidder's evaluation, including mandatory/optional criteria and detailed reasoning.", small_style))
    content.append(Spacer(1, 6 * mm))

    for b in bidder_results:
        b_name = b.get("bidder_name", "Unknown Bidder")
        b_verdict = b.get("verdict", "UNKNOWN").replace("_", " ")
        v_color = PASS_COLOR if b_verdict == "ELIGIBLE" else (FAIL_COLOR if b_verdict == "NOT ELIGIBLE" else REVIEW_COLOR)
        
        b_title_style = ParagraphStyle('BTitle', parent=styles['Heading3'], fontSize=12, textColor=ACCENT_BLUE, spaceAfter=2)
        b_verdict_style = ParagraphStyle('BVerdict', parent=styles['Normal'], fontSize=10, textColor=v_color, spaceAfter=8, fontName='Helvetica-Bold')
        
        content.append(HRFlowable(width="100%", color=colors.HexColor("#e2e8f0"), thickness=1))
        content.append(Spacer(1, 4 * mm))
        content.append(Paragraph(f"Bidder: {b_name}", b_title_style))
        content.append(Paragraph(f"Final Verdict: {b_verdict}", b_verdict_style))

        # Criteria details for this bidder
        b_evals = b.get("evaluation", [])
        if b_evals:
            for e in b_evals:
                cid = e.get("criterion_id", "")
                c_name = e.get("criteria_name", cid)
                result = e.get("result", "")
                reason = e.get("reason", "No reason provided.")
                
                # Identify if mandatory from global criteria
                criterion = next((c for c in criteria if c.get("criterion_id") == cid), {})
                c_type = "Mandatory" if criterion.get("mandatory", True) else "Optional"
                
                status_text = "[PASS]" if result == "PASS" else ("[FAIL]" if result == "FAIL" else "[REVIEW]")
                content.append(Paragraph(
                    f"<b>{cid} - {c_name}</b> [{c_type}] {status_text}", body_style
                ))
                content.append(Paragraph(f"→ {reason}", small_style))
                content.append(Spacer(1, 2 * mm))
        else:
            content.append(Paragraph("No detailed evaluation data available.", small_style))
        
        content.append(Spacer(1, 6 * mm))

    # ── Sign-Off ──
    content.append(Spacer(1, 12 * mm))
    content.append(HRFlowable(width="100%", color=ACCENT_BLUE, thickness=1.5))
    content.append(Paragraph("Authorized Sign-Off", section_style))
    signoff_style = ParagraphStyle(
        'ConsoSignOff', parent=styles['Normal'],
        fontSize=10, textColor=colors.black, spaceAfter=6, leading=18
    )
    content.append(Paragraph(
        "I have reviewed the above consolidated evaluation report and confirm the findings.", signoff_style
    ))
    content.append(Spacer(1, 6 * mm))

    signoff_data = [
        ["Procurement Officer Name:", "______________________________"],
        ["Designation:", "______________________________"],
        ["Signature:", "______________________________"],
        ["Date:", "______________________________"],
        ["Audit Hash Verified:", sha256[:32] + "..."],
    ]
    signoff_table = Table(signoff_data, colWidths=[150, 310])
    signoff_table.setStyle(TableStyle([
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('TEXTCOLOR', (0, 0), (0, -1), ACCENT_BLUE),
        ('TEXTCOLOR', (1, -1), (1, -1), TEXT_GRAY),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('LINEBELOW', (0, 0), (-1, -2), 0.5, colors.HexColor("#e2e8f0")),
    ]))
    content.append(signoff_table)

    # ── Footer ──
    content.append(Spacer(1, 10 * mm))
    content.append(HRFlowable(width="100%", color=TEXT_GRAY, thickness=0.5))
    content.append(Paragraph(
        f"Generated by InferX v2.0 | {timestamp} | SHA-256: {sha256[:16]}...",
        ParagraphStyle('ConsoFooter', parent=styles['Normal'], fontSize=9,
                       textColor=TEXT_GRAY, alignment=TA_CENTER)
    ))
    content.append(Paragraph(
        "This is a machine-generated report. All evaluations require human verification before final decisions.",
        ParagraphStyle('ConsoDisclaimer', parent=styles['Normal'], fontSize=9,
                       textColor=REVIEW_COLOR, alignment=TA_CENTER, spaceBefore=4)
    ))

    doc.build(content)
    return output_path

